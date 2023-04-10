from tkinter import*
from tkinter.ttk import Separator, Style
from os import startfile
from os.path import basename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from tkinter.filedialog import askopenfiles, asksaveasfilename
from ctypes import windll
from threading import Thread
from time import sleep

windll.shcore.SetProcessDpiAwareness(1)
win = Tk()
win.title('Py-Merger GUI by Flo')
win.iconbitmap('Images/pymerger.ico')
win.geometry(f'750x500+{round(win.winfo_screenwidth()/4)}+{round(win.winfo_screenheight()/4)}')
win.minsize(650, 458)

docs = {'btn_merge': 'Merge selected sheets.', 'btn_import': 'Import Excel Files.', 'btn_bin':'Delete selected items.', 'btn_add':'Add to merge list.', 'btn_base':'Set sheet as base.', 'btn_play':'Start selected workbook.', 'btn_find':'Locate in explorer.'}
workbooks = {}
worksheets = {}
wb_path = {}
selection_keys = []
base = 0

"""
def manage_data():
	while 1:
		wb = list(workbooks.keys())
		ws = list(worksheets.keys())
		upt = selection_keys
		bs = base
		sleep(1)
		if wb != list(workbooks.keys()) or ws != list(worksheets.keys()) or upt != selection_keys or bs != base:
			print(f'Workbooks: {list(workbooks.keys())}')
			print(f'Worksheets: {list(worksheets.keys())}')
			print(f'Updates: {selection_keys}')
			print(f'Base: {base}\n')
		
Thread(target=manage_data, daemon=True).start()
"""

def start_wb(*args):
	active = workbooks_list.curselection()
	if active != ():
		fn = wb_path[workbooks_list.get(active[0])]
		if args == ():
			startfile(fn)
		else:
			startfile(fn.replace(basename(fn), ''))

def import_wb(*args):
	global base
	if args != ():
		filenames = list(args)
	else:
		filenames = [elem.name for elem in askopenfiles(filetypes=(("Excel File","*.xlsx"),))]
	for filename in filenames:
		alias_name = basename(filename)
		wb_path[alias_name] = filename
		workbooks[alias_name] = load_workbook(filename) # workbooks = {'workbook_name':Workbook}
		if not alias_name in workbooks_list.get('0', 'end'):
			workbooks_list.insert('end', alias_name)
	for workbook_name in list(workbooks.keys()):
		x = 0
		for sheet in enumerate(workbooks[workbook_name].sheetnames):
			alias_name = f'{sheet[1]} ({workbook_name})'
			worksheets[alias_name] = workbooks[workbook_name].worksheets[x] # worksheets = {'worksheet_name-workbook_name':worksheet}
			if not alias_name in worksheets_list.get('0', 'end'):
				worksheets_list.insert('end', alias_name)
			if not base:
				base = alias_name
				worksheets_list.itemconfig(0, bg='#3498db', selectbackground='#2980b9')
			x += 1

def delete_item():
	global base
	global selection_keys
	wb_idx = workbooks_list.curselection()
	ws_idx = worksheets_list.curselection()
	for wb in wb_idx:
		temp = [elem if elem.endswith(workbooks_list.get(wb)+')') else 0 for elem in worksheets_list.get('0', 'end')]
		del workbooks[workbooks_list.get(wb)]
		workbooks_list.delete(wb)
		for elem in temp: # Pour toutes les feuilles associées
			if elem != 0:
				if elem in selection_keys:
					selection_keys.remove(elem)
				del worksheets[elem]
				worksheets_list.delete(worksheets_list.get('0', 'end').index(elem))
	for ws in ws_idx:
		item = worksheets_list.get(ws)
		if item in selection_keys:
			selection_keys.remove(item)
		del worksheets[item]
		worksheets_list.delete(ws)
	lb_ws = worksheets_list.get('0','end')

	if not base in lb_ws: # Base supprimée
		if len(lb_ws)>0:
			base = lb_ws[0]
			worksheets_list.itemconfig(0, bg='#3498db', selectbackground='#2980b9')
			if base in selection_keys:
				selection_keys.remove(base)
		else:
			base = 0

def add_selection():
	global selection_keys
	idx = worksheets_list.curselection()
	if idx != ():
		value = worksheets_list.get(idx)
		if value == base:
			pass
			#print("You can't set base as update of himself.")
		elif value in selection_keys:
			selection_keys.remove(value)
			worksheets_list.itemconfig(idx, bg='SystemButtonFace', selectbackground='#D3D3D3')
		else:
			selection_keys.append(value)
			worksheets_list.itemconfig(idx, bg='#2ecc71', selectbackground='#27ae60')
		#print(selection_keys)
	else:
		pass
		#print('Select at least one sheet.')

def set_base():
	global base
	global selection_keys
	idx = worksheets_list.curselection()
	if idx == ():
		pass
		#print('Select at least one sheet.')
	else:
		idx = idx[0]
		if base in worksheets_list.get('0', 'end'):
			temp_idx = worksheets_list.get('0', 'end').index(base)
			worksheets_list.itemconfig(temp_idx, bg='SystemButtonFace', selectbackground='#D3D3D3')
		base = worksheets_list.get(idx)
		worksheets_list.itemconfig(idx, bg='#3498db', selectbackground='#2980b9')
		if base in selection_keys:
			selection_keys.remove(base)

def merge_internal_func(old, upt, new_ws):
	alphabet = [chr(k) for k in range(65, 91)]
	reorder_rows = lambda sheet: {elem[0]:list(elem[1:]) for elem in list(sheet.values)[1:]} # {A1:[B1, C1...], A2:[B2, C2...]...}
	reorder_col_names = lambda sheet: {list(sheet.values)[0][k+1]:k for k in range(len(list(sheet.values)[0])-1)}
	sheet_origin = lambda sheet: list(sheet.values)[0][0]
	old_rows = reorder_rows(old)
	old_titles = reorder_col_names(old)
	upt_rows = reorder_rows(upt)
	upt_titles = reorder_col_names(upt)
	new_titles = list(old_titles.keys()) + list(set(upt_titles.keys()) - set(old_titles.keys()))
	new_titles = {new_titles[k]:k for k in range(len(new_titles))} # anciens titres + les nouveaux
	for elem in old_rows:
		while len(old_rows[elem])<len(new_titles):
			old_rows[elem].append(None)
	for elem in upt_rows:
		if elem in old_rows:
			for title in upt_titles:
				if isinstance(old_rows[elem][new_titles[title]], int):
					if old_rows[elem][new_titles[title]] > upt_rows[elem][upt_titles[title]]:
						old_rows[elem][new_titles[title]] = (upt_rows[elem][upt_titles[title]], 0)
					elif old_rows[elem][new_titles[title]] < upt_rows[elem][upt_titles[title]]:
						old_rows[elem][new_titles[title]] = (upt_rows[elem][upt_titles[title]], 1)
					else:
						old_rows[elem][new_titles[title]] = upt_rows[elem][upt_titles[title]]
				else:
					old_rows[elem][new_titles[title]] = upt_rows[elem][upt_titles[title]]
		else:
			old_rows[elem] = [None for k in range(len(new_titles))]
			for title in upt_titles:
				old_rows[elem][new_titles[title]] = upt_rows[elem][upt_titles[title]]
	# On écrit brut data 0 (Operationnel)
	for elem in enumerate([sheet_origin(old)] + list(new_titles.keys())): # Ligne des titres
		index = alphabet[elem[0]] + '1'
		new_ws[index] = elem[1]
		new_ws[index].font = Font(bold=True)
	for idx, key in enumerate(old_rows, 2): 
		index = 'A' + str(idx)
		new_ws[index] = key
		for i in range(len(old_rows[key])): 
			index = alphabet[i+1] + str(idx)
			if isinstance(old_rows[key][i], tuple):
				col = old_rows[key][i][1]
				new_ws[index] = old_rows[key][i][0]
				new_ws[index].font = Font(color="2ecc71" if col else 'e74c3c')
			else:
				new_ws[index] = old_rows[key][i]

def merge_action():
	if not base or len(selection_keys)==0:
		return ''
	new_wb = Workbook()
	new_ws = new_wb.active
	new_ws.title= 'py_merged'
	base_ws = worksheets[base]
	upt_ws = worksheets[selection_keys[0]]
	merge_internal_func(base_ws, upt_ws, new_ws)
	for ws in selection_keys[1:]:
		upt_ws = worksheets[ws]
		merge_internal_func(new_ws, upt_ws, new_ws)
	new_fn = asksaveasfilename(filetypes=(("Excel File","*.xlsx"),))
	if new_fn != '':
		new_fn = new_fn + '.xlsx' if not new_fn.endswith('.xlsx') else new_fn
		new_wb.save(new_fn)
		import_wb(new_fn)    

for k in range(10):
	win.rowconfigure(k, weight=1)
	win.columnconfigure(k, weight=1)

def stop_hover(event):
	title_var.set('Py-Merger panel')
	my_title.config(font=('bahnschrift bold', 22), image='')

def hover(data):
	title_var.set(data)
	my_title.config(font=('bahnschrift light', 22), image=pic_tip)

pic_tip = PhotoImage(file='Images/Tips.png').subsample(2)
title_var = StringVar(win, 'Py-Merger panel')
my_title = Label(master=win, textvariable=title_var, font=('bahnschrift bold', 22), compound='left')
my_title.grid(columnspan=10, sticky='w', padx=20)

bottom_frame = Frame(master=win)
bottom_frame.grid(row=9, column=0, columnspan=10, sticky='nsew')

for k in range(3):
	bottom_frame.rowconfigure(k, weight=1)
	bottom_frame.columnconfigure(k, weight=1)

pic1 = PhotoImage(file="Images/Pic 1.png")
btn_import = Button(master=bottom_frame, image=pic1, relief='flat', borderwidth=0, cursor='hand2', command=import_wb)
btn_import.grid(row=1, column=0, sticky='nse', ipady=10)
btn_import.bind('<Enter>', lambda event: hover(docs['btn_import']))
btn_import.bind('<Leave>', stop_hover)

pic2 = PhotoImage(file="Images/Pic 2.png")
btn_merge = Button(master=bottom_frame, command=merge_action, image=pic2, relief='flat', borderwidth=0, cursor='hand2')
btn_merge.grid(row=1, column=2, sticky='nsw', ipady=10)
btn_merge.bind('<Enter>', lambda event: hover(docs['btn_merge']))
btn_merge.bind('<Leave>', stop_hover)

interactive_part = Frame(master=win, borderwidth=0, highlightthickness=0)
interactive_part.grid(row=1, column=0, columnspan=10, rowspan=8, pady=15, sticky='nsew')

interactive_part.columnconfigure(0, weight = 5)
interactive_part.columnconfigure(1, weight = 5)
interactive_part.columnconfigure(2, weight = 1)

interactive_part.rowconfigure(0, weight = 1)
interactive_part.rowconfigure(1, weight = 10)

legend_lft = Label(master=interactive_part, text='Imported Workbooks', font=('arial gras', 12))
legend_lft.grid(row=0, column=0)

legend_rgt = Label(master=interactive_part, text='Imported Worksheets', font=('arial gras', 12))
legend_rgt.grid(row=0, column=1)

bin_pic = PhotoImage(file='Images/recycle_bin.png').subsample(1)
btn_bin = Button(master=interactive_part, command=delete_item ,image=bin_pic, relief='flat', borderwidth=0, cursor='hand2', activebackground='#7f8c8d')
btn_bin.grid(row=0, column=2, ipady=10, ipadx=5, sticky='nsew')
btn_bin.bind('<Enter>', lambda event: [hover(docs['btn_bin']), btn_bin.config(bg="#D3D3D3")])
btn_bin.bind('<Leave>', lambda event: [stop_hover(None), btn_bin.config(bg='SystemButtonFace')])

workbooks_list = Listbox(interactive_part, activestyle='none', selectforeground='black',  relief='flat', highlightthickness=0, bg='SystemButtonFace', justify='center', font=('bahnschrift light', 10), selectbackground='#D3D3D3')
workbooks_list.grid(row=1, column=0, sticky='nsew')

worksheets_list = Listbox(interactive_part, activestyle='none', selectforeground='black', relief='flat', highlightthickness=0, bg='SystemButtonFace', justify='center', font=('bahnschrift light', 10), selectbackground='#D3D3D3')
worksheets_list.grid(row=1, column=1, sticky='nsew')

btn_frame = Frame(master=interactive_part)
btn_frame.grid(row=1, column=2, sticky='nsew')
btn_frame.columnconfigure(0, weight=1)

add_pic = PhotoImage(file='Images/add_elem.png').subsample(1)
btn_add = Button(master=btn_frame, command=add_selection, relief='flat', image=add_pic, bd=0, cursor='hand2', activebackground='#7f8c8d')
btn_add.grid(sticky='nsew', ipady=10, ipadx=5)
btn_add.bind('<Enter>', lambda event: [hover(docs['btn_add']), btn_add.config(bg="#D3D3D3")])
btn_add.bind('<Leave>', lambda event: [stop_hover(None), btn_add.config(bg="SystemButtonFace")])

base_pic = PhotoImage(file='Images/base_ico.png').subsample(1)
btn_base = Button(master=btn_frame, command=set_base, relief='flat', image=base_pic, bd=0, cursor='hand2', activebackground='#7f8c8d')
btn_base.grid(sticky='nsew', ipady=10, ipadx=5)
btn_base.bind('<Enter>', lambda event: [hover(docs['btn_base']), btn_base.config(bg="#D3D3D3")])
btn_base.bind('<Leave>', lambda event: [stop_hover(None), btn_base.config(bg="SystemButtonFace")])

play_pic = PhotoImage(file='Images/play_icon.png').subsample(1)
btn_play = Button(master=btn_frame, command=start_wb, relief='flat', image=play_pic, bd=0, cursor='hand2', activebackground='#7f8c8d')
btn_play.grid(sticky='nsew', ipady=10, ipadx=5)
btn_play.bind('<Enter>', lambda event: [hover(docs['btn_play']), btn_play.config(bg="#D3D3D3")])
btn_play.bind('<Leave>', lambda event: [stop_hover(None), btn_play.config(bg="SystemButtonFace")])

folder_pic = PhotoImage(file='Images/folder_pic.png').subsample(1)
btn_find = Button(master=btn_frame, command=lambda: start_wb(None), relief='flat', image=folder_pic, bd=0, cursor='hand2', activebackground='#7f8c8d')
btn_find.grid(sticky='nsew', ipady=10, ipadx=5)
btn_find.bind('<Enter>', lambda event: [hover(docs['btn_find']), btn_find.config(bg="#D3D3D3")])
btn_find.bind('<Leave>', lambda event: [stop_hover(None), btn_find.config(bg="SystemButtonFace")])

Separator(interactive_part, orient='vertical').grid(column=0, row=0, rowspan=2, sticky='nse')
Separator(interactive_part, orient='vertical').grid(column=2, row=0, rowspan=2, sticky='nsw')
Separator(interactive_part, orient='horizontal').grid(column=0, row=0, columnspan=3, sticky='new')
Separator(interactive_part, orient='horizontal').grid(column=0, row=0, columnspan=3, sticky='sew')
Separator(interactive_part, orient='horizontal').grid(column=0, row=1, columnspan=3, sticky='sew')

win.mainloop()